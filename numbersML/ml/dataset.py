"""
PyTorch Dataset for loading wide_vectors + target_values.

Data flow:
  wide_vectors were calculated and stored in DB.
  candles_1s.target_value (JSONB) contains 'normalized_value' (0 to 1).
  We load this pre-calculated target directly.

The dataset handles:
  - Loading data from PostgreSQL
  - Loading pre-calculated normalized targets (0-1) from DB
  - Aligning X (wide vectors) with Y (future normalized value)
  - Feature normalization (StandardScaler)
  - Train/val/test splitting by time
"""

import json
import logging
import os
from datetime import datetime, timedelta, timezone
from typing import List, Optional, Tuple

import numpy as np
import torch
from torch.utils.data import Dataset, DataLoader
import psycopg2

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

from ml.config import DatabaseConfig, DataConfig

logger = logging.getLogger(__name__)


class WideVectorDataset(Dataset):
    """
    Dataset that loads wide_vectors and corresponding target_values.

    Each sample is a sequence of `sequence_length` consecutive wide_vectors,
    with the target being the target_value for a SPECIFIC SYMBOL at the end of the sequence.

    Important:
    - wide_vector contains ALL symbols' data at timestamp T
    - target_value is for ONE specific symbol (e.g., BTC/USDC)
    - We train a model to predict THAT symbol's target from the wide vector
    """

    def __init__(
        self,
        db_config: DatabaseConfig,
        data_config: DataConfig,
        start_time: datetime,
        end_time: datetime,
        mean: Optional[np.ndarray] = None,
        std: Optional[np.ndarray] = None,
        feature_mask: Optional[np.ndarray] = None,
        sequence_length: int = 60,
    ):
        self.db_config = db_config
        self.data_config = data_config
        self.start_time = start_time
        self.end_time = end_time
        self.sequence_length = sequence_length
        self.mean = mean
        self.std = std
        self.feature_mask = feature_mask
        self.target_symbol = data_config.target_symbol

        # Load data
        self.vectors, self.targets, self.timestamps = self._load_data()

        # Compute normalization params if not provided
        if self.mean is None or self.std is None:
            all_vectors = np.vstack(self.vectors)
            
            # Apply StandardScaler: zero mean, unit variance
            self.mean = np.mean(all_vectors, axis=0)
            self.std = np.std(all_vectors, axis=0)
            
            # Add epsilon to prevent division by zero
            epsilon = 1e-8
            self.std = np.where(self.std < epsilon, epsilon, self.std)

            # Filter out low-variance features to prevent overfitting
            # Higher threshold = fewer features = less overfitting
            min_std = 0.05  # Keep only top features with meaningful variance
            self.feature_mask = self.std > min_std
            
            # Apply feature mask
            self.mean = self.mean[self.feature_mask]
            self.std = self.std[self.feature_mask]
            
            # Normalize vectors: x = (x - mean) / std
            self.vectors = [(v[self.feature_mask] - self.mean) / self.std for v in self.vectors]
            
            print(f'Feature normalization: StandardScaler (zero mean, unit variance)')
            print(f'  Original features: {all_vectors.shape[1]}')
            print(f'  Features after mask: {len(self.mean)} (removed {all_vectors.shape[1] - len(self.mean)} low-variance)')
        else:
            # Apply existing feature mask and normalization
            self.vectors = [(v[self.feature_mask] - self.mean) / self.std for v in self.vectors]

        # Build valid sequence indices
        self._build_sequences()

    def _validate_temporal_consistency(self, timestamps: List[datetime], vectors: List[np.ndarray]):
        """
        DATA QUALITY VALIDATION:
        Ensures that:
        1. Each timestamp has exactly one wide_vector (no duplicates)
        2. Timestamps grow by exactly +1 second (consecutive 1-second intervals)
        
        Logs warnings if any violations are found (does not break training).
        """
        if len(timestamps) == 0:
            return
        
        logger.info(f"Validating temporal consistency for {len(timestamps)} samples...")
        
        # Check 1: Ensure no duplicate timestamps (each time has exactly one wide_vector)
        seen_timestamps = {}
        duplicate_count = 0
        first_duplicate_ts = None
        for i, ts in enumerate(timestamps):
            if ts in seen_timestamps:
                duplicate_count += 1
                if first_duplicate_ts is None:
                    first_duplicate_ts = ts
                if duplicate_count <= 5:  # Log first 5 duplicates for debugging
                    logger.warning(
                        f"⚠ DUPLICATE TIMESTAMP: {ts} appears at indices {seen_timestamps[ts]} and {i}"
                    )
            else:
                seen_timestamps[ts] = i
        
        if duplicate_count > 0:
            logger.warning(
                f"⚠ DATA QUALITY WARNING: Found {duplicate_count} duplicate timestamps "
                f"(first duplicate: {first_duplicate_ts}). "
                f"Each time should have exactly one wide_vector. Training will continue but results may be affected."
            )
        
        # Check 2: Ensure timestamps grow by exactly +1 second
        gap_violations = []
        max_violations_to_log = 10
        
        for i in range(1, len(timestamps)):
            prev_ts = timestamps[i - 1]
            curr_ts = timestamps[i]
            time_diff = (curr_ts - prev_ts).total_seconds()
            
            if time_diff != 1.0:
                gap_violations.append({
                    'index': i,
                    'prev_time': prev_ts,
                    'curr_time': curr_ts,
                    'diff_seconds': time_diff
                })
                
                if len(gap_violations) <= max_violations_to_log:
                    logger.warning(
                        f"⚠ TEMPORAL GAP at index {i}: "
                        f"{prev_ts} -> {curr_ts} (diff={time_diff}s, expected=1.0s)"
                    )
        
        if gap_violations:
            # Summarize all violations
            diff_counts = {}
            for v in gap_violations:
                diff = v['diff_seconds']
                diff_counts[diff] = diff_counts.get(diff, 0) + 1
            
            violation_summary = ", ".join([
                f"{diff}s: {count} times" for diff, count in sorted(diff_counts.items())
            ])
            
            logger.warning(
                f"⚠ DATA QUALITY WARNING: Found {len(gap_violations)} temporal gaps not exactly +1 second. "
                f"Violations: {violation_summary}. "
                f"Database state may be inconsistent. Training will continue but results may be affected."
            )
        
        if duplicate_count == 0 and len(gap_violations) == 0:
            logger.info("✓ Temporal consistency validated: All timestamps are unique and grow by exactly +1 second")

    def _load_data(
        self,
    ) -> Tuple[List[np.ndarray], List[float], List[datetime]]:
        """Load wide_vectors and targets directly from DB."""
        conn = psycopg2.connect(
            host=self.db_config.host,
            port=self.db_config.port,
            dbname=self.db_config.dbname,
            user=self.db_config.user,
            password=self.db_config.password,
        )

        try:
            # First, get the symbol_id for the target symbol
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id FROM symbols WHERE symbol = %s",
                    (self.target_symbol,)
                )
                symbol_row = cur.fetchone()
                if not symbol_row:
                    raise ValueError(f"Symbol '{self.target_symbol}' not found in database")
                symbol_id = symbol_row[0]

            # Now load data with a named cursor
            with conn.cursor(name="ml_data_cursor") as cur:
                cur.itersize = 5000

                # Load wide_vectors with close price and target
                # Target will be converted to relative change: (target[t+h] - close[t]) / close[t]
                query = """
                    SELECT
                        wv.time,
                        wv.vector,
                        wv.vector_size,
                        (c.target_value->>'normalized_value')::float AS target,
                        c.close
                    FROM wide_vectors wv
                    JOIN candles_1s c ON c.time = wv.time AND c.symbol_id = %s
                    WHERE wv.time >= %s AND wv.time < %s
                      AND wv.vector_size >= 50
                      AND c.target_value IS NOT NULL
                      AND (c.target_value->>'normalized_value') IS NOT NULL
                      AND c.close IS NOT NULL
                    ORDER BY wv.time
                """
                cur.execute(query, (symbol_id, self.start_time, self.end_time))

                vectors = []
                targets = []
                closes = []
                timestamps = []
                prev_size = None

                while True:
                    batch = cur.fetchmany(5000)
                    if not batch:
                        break

                    for row in batch:
                        ts, vector_json, vec_size, target, close = row

                        # Parse vector from JSONB
                        if isinstance(vector_json, str):
                            vec = np.array(json.loads(vector_json), dtype=np.float32)
                        else:
                            vec = np.array(vector_json, dtype=np.float32)

                        # Handle variable-length vectors by padding
                        if prev_size is None:
                            prev_size = len(vec)
                        elif len(vec) != prev_size:
                            # Pad shorter vectors with zeros
                            if len(vec) < prev_size:
                                vec = np.pad(vec, (0, prev_size - len(vec)))
                            else:
                                vec = vec[:prev_size]

                        vectors.append(vec)
                        targets.append(float(target))
                        closes.append(float(close))
                        timestamps.append(ts)

        finally:
            conn.close()

        # DATA QUALITY VALIDATION: Ensure exact +1 second intervals and unique timestamps
        self._validate_temporal_consistency(timestamps, vectors)

        # Convert absolute targets to relative changes
        # relative_target[i] = (target[t+i+horizon] - close[t+i]) / close[t+i]
        # This makes the target scale-invariant and learnable
        horizon = self.data_config.prediction_horizon
        if len(targets) > horizon:
            # Apply horizon shift
            shifted_targets = targets[horizon:]
            shifted_vectors = vectors[:-horizon]
            shifted_timestamps = timestamps[:-horizon]
            shifted_closes = closes[:-horizon]
            
            # Compute relative change: (future_target - current_close) / current_close
            relative_targets = [
                (ft - c) / c if c != 0 else 0.0
                for ft, c in zip(shifted_targets, shifted_closes)
            ]
            
            targets = relative_targets
            vectors = shifted_vectors
            timestamps = shifted_timestamps
        else:
            raise ValueError(f"Not enough data for prediction horizon {horizon}")

        print(f'Loaded {len(vectors)} samples, {len(targets)} targets, {len(timestamps)} timestamps')
        print(f'  Symbol: {self.target_symbol}, target id: {symbol_id}')
        print(f'  Time range: {timestamps[0] if timestamps else "N/A"} to {timestamps[-1] if timestamps else "N/A"}')
        print(f'  Prediction horizon: {horizon}s (relative change from close)')
        print(f'  Target stats: mean={np.mean(targets):.6f}, std={np.std(targets):.6f}')

        if len(vectors) < self.data_config.min_samples:
            raise ValueError(
                f"Insufficient samples: {len(vectors)} < {self.data_config.min_samples}"
            )

        return vectors, targets, timestamps

    def _build_sequences(self):
        """Build valid sequence start indices."""
        self.sequence_indices = []
        n = len(self.vectors)

        for i in range(n - self.sequence_length + 1):
            # Check if timestamps are consecutive (allow 1-2 second gaps)
            start_ts = self.timestamps[i]
            end_ts = self.timestamps[i + self.sequence_length - 1]
            expected_duration = timedelta(seconds=self.sequence_length - 1)
            actual_duration = end_ts - start_ts

            # Allow small gaps (up to 5 seconds)
            if actual_duration <= expected_duration + timedelta(seconds=5):
                self.sequence_indices.append(i)

    def __len__(self) -> int:
        return len(self.sequence_indices)

    def __getitem__(self, idx: int) -> Tuple[torch.Tensor, torch.Tensor]:
        start_idx = self.sequence_indices[idx]
        end_idx = start_idx + self.sequence_length

        # Stack vectors into (sequence_length, features)
        sequence = np.stack(self.vectors[start_idx:end_idx])

        # Target is the target_value at the end of the sequence
        target = self.targets[end_idx - 1]

        return torch.from_numpy(sequence), torch.tensor(target, dtype=torch.float32)

    def get_feature_dim(self) -> int:
        """Get the feature dimension of the vectors."""
        return len(self.vectors[0]) if self.vectors else 0


def export_training_data_to_excel(
    db_config: DatabaseConfig,
    data_config: DataConfig,
    start_time: datetime,
    end_time: datetime,
    output_path: str = "ml/data_export/training_data.xlsx",
    max_rows: int = 3000,
):
    """
    Export raw training data (wide vectors + target values) to Excel file.
    Creates one row per timestamp with all features as columns plus target value.
    
    Useful for manual inspection, feature analysis, and debugging.
    """
    if not HAS_EXCEL:
        logger.warning("openpyxl not available, skipping Excel export")
        return
    
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    logger.info(f"Exporting training data to Excel: {output_path}")
    print(f"\nExporting first {max_rows} rows of training data to {output_path}...")
    
    conn = psycopg2.connect(
        host=db_config.host,
        port=db_config.port,
        dbname=db_config.dbname,
        user=db_config.user,
        password=db_config.password,
    )
    
    try:
        # Get symbol_id
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id FROM symbols WHERE symbol = %s",
                (data_config.target_symbol,)
            )
            symbol_row = cur.fetchone()
            symbol_id = symbol_row[0] if symbol_row else None
        
        # Load column_names from first wide_vector
        with conn.cursor() as cur:
            cur.execute(
                "SELECT column_names FROM wide_vectors WHERE time >= %s AND time < %s AND vector_size >= 50 LIMIT 1",
                (start_time, end_time)
            )
            col_row = cur.fetchone()
            column_names = col_row[0] if col_row else []
        
        n_features = len(column_names)
        
        # Load data
        with conn.cursor(name="export_cursor") as cur:
            cur.itersize = 5000
            query = """
                SELECT
                    wv.time,
                    wv.vector,
                    (c.target_value->>'normalized_value')::float AS target,
                    c.close
                FROM wide_vectors wv
                JOIN candles_1s c ON c.time = wv.time AND c.symbol_id = %s
                WHERE wv.time >= %s AND wv.time < %s
                  AND wv.vector_size >= 50
                  AND c.target_value IS NOT NULL
                  AND (c.target_value->>'normalized_value') IS NOT NULL
                  AND c.close IS NOT NULL
                ORDER BY wv.time
            """
            cur.execute(query, (symbol_id, start_time, end_time))

            rows = []
            closes = []
            row_count = 0
            while row_count < max_rows:
                batch = cur.fetchmany(min(5000, max_rows - row_count))
                if not batch:
                    break
                for row in batch:
                    ts, vector_json, target, close = row
                    # Strip timezone for Excel compatibility
                    if ts.tzinfo is not None:
                        ts = ts.replace(tzinfo=None)
                    # Parse vector
                    if isinstance(vector_json, str):
                        vec = json.loads(vector_json)
                    else:
                        vec = list(vector_json)
                    # Pad/truncate to match column_names
                    if len(vec) < n_features:
                        vec = vec + [0.0] * (n_features - len(vec))
                    elif len(vec) > n_features:
                        vec = vec[:n_features]
                    rows.append([ts] + vec + [target, close])
                    closes.append(float(close))
                    row_count += 1
                    if row_count >= max_rows:
                        break
    finally:
        conn.close()
    
    # Apply prediction_horizon shift (same as training)
    # row[i] = [time, vec..., target, close]
    # After shift: vector at time t → target at time t+horizon
    # Compute relative target: (target[t+h] - close[t]) / close[t]
    horizon = data_config.prediction_horizon
    if len(rows) > horizon * 2:
        n = len(rows)
        n_out = n - horizon
        
        shifted_rows = []
        for i in range(n_out):
            time_val = rows[i][0]
            vec_data = rows[i][1:-2]  # everything between time and target/close
            future_target = rows[i + horizon][-2]  # target from row[i+horizon]
            current_close = rows[i][-1]  # close at time t
            
            # Relative change
            if current_close != 0:
                relative_target = (future_target - current_close) / current_close
            else:
                relative_target = 0.0
            
            shifted_rows.append([time_val] + vec_data + [relative_target])
        
        rows = shifted_rows[:max_rows]
    else:
        print("  Not enough data for prediction horizon, skipping export")
        return
    
    if not rows:
        print("  No data to export")
        return
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Training Data"
    
    # Add info row above headers
    ws['A1'] = f"Training data for {data_config.target_symbol}"
    ws['A2'] = f"Prediction horizon: +{horizon}s | relative target = (future_target - close) / close"
    ws['A3'] = f"Each row: wide_vector at time t with relative target from time t+{horizon}s (percentage)"
    for r in range(1, 4):
        ws.cell(row=r, column=1).font = Font(italic=True, size=9)
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", wrap_text=True)
    
    # Write headers (start at row 5)
    headers = ["time"] + list(column_names) + [f"rel_target (t+{horizon}s, %)"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data (start at row 6)
    for row_idx, row_data in enumerate(rows, 6):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Column widths
    ws.column_dimensions['A'].width = 25  # Time column
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    
    # Freeze header row
    ws.freeze_panes = 'A6'
    
    # Auto-filter
    ws.auto_filter.ref = f"A5:{get_column_letter(len(headers))}{len(rows) + 5}"
    
    wb.save(output_path)
    print(f"  ✓ Exported {len(rows)} rows × {len(headers)} columns to {output_path}")
    logger.info(f"Exported {len(rows)} rows to {output_path}")


def create_data_loaders(
    db_config: DatabaseConfig,
    data_config: DataConfig,
) -> Tuple[DataLoader, DataLoader, DataLoader, np.ndarray, np.ndarray, np.ndarray]:
    """
    Create train, validation, and test data loaders.

    Returns:
        train_loader, val_loader, test_loader, mean, std, feature_mask
    """
    # First, check what data is available
    conn = psycopg2.connect(
        host=db_config.host,
        port=db_config.port,
        dbname=db_config.dbname,
        user=db_config.user,
        password=db_config.password,
    )
    
    try:
        with conn.cursor() as cur:
            # Get the actual data range available
            cur.execute("""
                SELECT 
                    MIN(wv.time) as earliest,
                    MAX(wv.time) as latest
                FROM wide_vectors wv
                INNER JOIN candles_1s c ON c.time = wv.time AND c.symbol_id = (
                    SELECT id FROM symbols WHERE symbol = %s
                )
                WHERE c.close IS NOT NULL
            """, (data_config.target_symbol,))
            row = cur.fetchone()
            if row and row[0] and row[1]:
                data_start = row[0]
                data_end = row[1]
            else:
                raise ValueError("No data available with target values")
    finally:
        conn.close()
    
    # Calculate time ranges based on available data and user request
    data_hours = (data_end - data_start).total_seconds() / 3600
    logger.info(f"Available data: {data_start} to {data_end} ({data_hours:.2f} hours)")

    # Check if user requested a specific time window (train_hours)
    # If train_hours is set and smaller than available data, we respect it
    # by shifting the end time to 'now' (or data_end) and start time backwards.
    
    # Default split: 70% train, 15% val, 15% test
    train_frac, val_frac, test_frac = 0.7, 0.15, 0.15

    # Determine the effective time window
    # If train_hours is defined in config (and > 0), we use it to limit the range
    requested_hours = data_config.train_hours
    
    if requested_hours > 0:
        # Use the requested hours starting from the latest data point
        total_requested_seconds = requested_hours * 3600
        
        # Ensure we don't go beyond available data
        effective_start = max(data_start, data_end - timedelta(seconds=total_requested_seconds))
        
        # If the requested window is larger than available data, use all data
        if (data_end - effective_start).total_seconds() <= 0:
            effective_start = data_start
            
        total_seconds = (data_end - effective_start).total_seconds()
        logger.info(f"Using requested window: {requested_hours} hours (Effective: {total_seconds/3600:.2f}h)")
    else:
        # Use all available data
        total_seconds = (data_end - data_start).total_seconds()
        effective_start = data_start
        logger.info(f"Using all available data: {total_seconds/3600:.2f} hours")

    # Calculate split durations
    train_seconds = total_seconds * train_frac
    val_seconds = total_seconds * val_frac
    test_seconds = total_seconds * test_frac

    # Time ranges (going backwards from data_end)
    test_end = data_end
    test_start = data_end - timedelta(seconds=test_seconds)

    val_end = test_start
    val_start = val_end - timedelta(seconds=val_seconds)

    train_end = val_start
    train_start = effective_start  # Start from the effective beginning

    logger.info(f"Train: {train_start} to {train_end}")
    logger.info(f"Val:   {val_start} to {val_end}")
    logger.info(f"Test:  {test_start} to {test_end}")

    # Load training data first to compute normalization
    train_dataset = WideVectorDataset(
        db_config=db_config,
        data_config=data_config,
        start_time=train_start,
        end_time=train_end,
        sequence_length=data_config.sequence_length,
    )

    mean = train_dataset.mean
    std = train_dataset.std
    feature_mask = train_dataset.feature_mask

    # Export training data to Excel (first 3000 rows)
    export_training_data_to_excel(
        db_config=db_config,
        data_config=data_config,
        start_time=train_start,
        end_time=train_end,
        output_path=os.path.join("ml", "data_export", "training_data.xlsx"),
        max_rows=3000,
    )

    # Load val and test with training normalization and feature mask
    val_dataset = WideVectorDataset(
        db_config=db_config,
        data_config=data_config,
        start_time=val_start,
        end_time=val_end,
        mean=mean,
        std=std,
        feature_mask=feature_mask,
        sequence_length=data_config.sequence_length,
    )

    test_dataset = WideVectorDataset(
        db_config=db_config,
        data_config=data_config,
        start_time=test_start,
        end_time=test_end,
        mean=mean,
        std=std,
        feature_mask=feature_mask,
        sequence_length=data_config.sequence_length,
    )

    train_loader = DataLoader(
        train_dataset,
        batch_size=data_config.batch_size,
        shuffle=True,
        num_workers=data_config.num_workers,
        pin_memory=False,
    )

    val_loader = DataLoader(
        val_dataset,
        batch_size=data_config.batch_size,
        shuffle=False,
        num_workers=data_config.num_workers,
        pin_memory=False,
    )

    test_loader = DataLoader(
        test_dataset,
        batch_size=data_config.batch_size,
        shuffle=False,
        num_workers=data_config.num_workers,
        pin_memory=False,
    )

    logger.info(f"Train samples: {len(train_dataset)}")
    logger.info(f"Val samples:   {len(val_dataset)}")
    logger.info(f"Test samples:  {len(test_dataset)}")
    logger.info(f"Feature dim:   {train_dataset.get_feature_dim()}")

    return train_loader, val_loader, test_loader, mean, std, feature_mask
